import sys
import os
import openpyxl

def rebrand_file(src_path, dest_path):
    print(f"Processing: {src_path} -> {dest_path}")
    if not os.path.exists(src_path):
        print(f"Error: source file {src_path} does not exist.")
        return False
    
    wb = openpyxl.load_workbook(src_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Iterate over all rows and cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value
                    
                    # Substring replacements
                    # 1. URL replacement
                    val = val.replace("https://thirulogasundar.github.io/CrowdSense", "https://thirulogasundar.github.io/ace-technologies/")
                    # 2. Case replacements
                    val = val.replace("CrowdSense", "Ace Technologies")
                    val = val.replace("crowdsense", "acetechnologies")
                    val = val.replace("CROWDSENSE", "ACE TECHNOLOGIES")
                    
                    cell.value = val

    wb.save(dest_path)
    print(f"Successfully saved to: {dest_path}")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python rebrand_excel.py <src_path> <dest_path>")
        sys.exit(1)
    
    src = sys.argv[1]
    dest = sys.argv[2]
    success = rebrand_file(src, dest)
    sys.exit(0 if success else 1)
